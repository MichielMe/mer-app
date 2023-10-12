import os
import time
import win32com.client as win32
import pythoncom
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from werkzeug.utils import secure_filename

def convert_xls_to_xlsx_with_excel(input_file):
    pythoncom.CoInitialize()  # Initialize the COM library
    output_file = input_file + 'x'
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    try:
        # Open the file in Excel
        wb = excel.Workbooks.Open(input_file)

        # Save the file in the new format
        wb.SaveAs(output_file, FileFormat=51)  # 51 represents the .xlsx format

        wb.Close()
    finally:
        # Ensure Excel is closed
        excel.Application.Quit()
        pythoncom.CoUninitialize()  # Uninitialize the COM library
    
    time.sleep(5)
    
    return output_file


def apply_excel_styles(filepath, form_data):
    # Load the workbook
    book = load_workbook(filepath)

    # Color mapping
    color_mapping = {
        "yellow": "FFFFCD34",
        "red": "FFDC2626",
        "blue": "FF60A5FA",
        "orange": "FFFB923C",
    }

    # Target texts and fill colors based on form data
    target_texts_colors = {
        "VRT": PatternFill(
            start_color=color_mapping[form_data['programma']],
            end_color=color_mapping[form_data['programma']],
            fill_type="solid",
        ),
        "KT_2023WEEK": PatternFill(
            start_color=color_mapping[form_data['wrap']],
            end_color=color_mapping[form_data['wrap']],
            fill_type="solid",
        ),
        "V00": PatternFill(
            start_color=color_mapping[form_data['ondertiteling']],
            end_color=color_mapping[form_data['ondertiteling']],
            fill_type="solid",
        ),
    }

    # Define the font style
    font_style = Font(bold=True)

    # Apply conditional formatting
    for sheet in book.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    for target_text, fill_style in target_texts_colors.items():
                        if target_text in str(cell.value):
                            cell.fill = fill_style
                            cell.font = font_style
                            break

    # Save the workbook
    book.save(filepath)
    
def save_uploaded_file(file, folder="uploads"):
    """
    Saves the uploaded file to the given folder and returns the filepath.
    If the folder doesn't exist, it gets created.

    Args:
    - file: The file to be saved.
    - folder: The folder to save the file in. Default is "uploads".

    Returns:
    - filepath: The path where the file was saved.
    """
    if not os.path.exists(folder):
        os.makedirs(folder)
    filename = secure_filename(file.filename)
    filepath = os.path.join(folder, filename)
    file.save(filepath)
    return filepath
