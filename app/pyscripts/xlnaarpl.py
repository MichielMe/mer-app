import time
import logging
import xml.etree.ElementTree as ET
from pathlib import Path, PurePath
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

def parse_excel_and_generate_playlist(filename):
    data_list = read_excel_to_list(filename)
    if not data_list:
        logging.error("Failed to read data from the Excel file or the file is empty.")
        return None
    
    output_filepath = generate_playlist(data_list, filename)
    return output_filepath


def read_excel_to_list(f):
    try:
        book = openpyxl.reader.excel.load_workbook(f)
    except (InvalidFileException, FileNotFoundError) as e:
        logging.error(str(e))
        return []
    
    sheet = book.active
    rows = sheet.max_row
    columns = 1  # Assuming we are only reading one column
    data = [sheet.cell(row=row, column=column).value for row in range(1, rows + 1) for column in range(1, columns + 1)]
    
    return data


def generate_playlist(data_list, filename):
    filename = PurePath(filename)
    ts = time.time()
    new_filename = filename.with_name(f"{filename.stem}_{ts}.sch")
    
    schedule_xml = generate_schedule_xml(data_list)
    
    try:
        tree = ET.ElementTree(schedule_xml)
        tree.write(open(new_filename, 'wb'), encoding='utf-8', xml_declaration=True)
        logging.info("Generated schedule with name: %s", new_filename)
        return new_filename
    except (IOError, TypeError) as e:
        logging.error("Error while saving the XML:", str(e))
        return None


def generate_schedule_xml(data_list):
    schedule = ET.Element('Schedule', Name='KetNet_31082018_0559_2055')
    events = ET.SubElement(schedule, 'Events', NotionalStartTime="31-08-2018 05:59:23:14", Channel="KetNet")
    
    for uid, entry in enumerate(data_list, start=1):
        event = ET.SubElement(events, 'Event', Uid=str(uid), FullyQualifiedType="KetNet Default Main Event")
        prev_id = ET.SubElement(event, 'PreviousUid')
        owner_id = ET.SubElement(event, 'OwnerUid')
        if uid == 1:
            prev_id.text = "-1"
            owner_id.text = "-1"
        else:
            prev_id.text = str(uid-1)
            owner_id.text = str(uid)
        is_fixed = ET.SubElement(event, 'IsFixed')
        is_fixed.text = "False"
        is_mediaball = ET.SubElement(event, 'IsMediaBall')
        is_mediaball.text = "False"
        fields = ET.SubElement(event, 'Fields')
        parameter = ET.SubElement(fields, 'Parameter', Name="EventName", Value=entry)
        parameter = ET.SubElement(fields, 'Parameter', Name="MaterialId", Value=entry)
        parameter = ET.SubElement(fields, 'Parameter', Name="PlayoutDeviceGroup", Value="Servers")
    
    return schedule
