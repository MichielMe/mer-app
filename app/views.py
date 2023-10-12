from flask import Blueprint, render_template, redirect, session, url_for, request, flash, jsonify, send_from_directory, send_file, current_app
from flask_login import login_user, logout_user, login_required, UserMixin, current_user
from app import login_manager
from flask_wtf import FlaskForm
from wtforms import StringField, PasswordField, SubmitField
from wtforms.validators import DataRequired
from .pyscripts.forms import UploadForm, ExcelUploadForm, ColorForm
from .pyscripts.xlnaarpl import parse_excel_and_generate_playlist
import pandas as pd
from .models import Material
import os
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import win32com.client as win32
import pythoncom
import time


main = Blueprint("main", __name__)

UPLOAD_FOLDER = '/uploads'
ALLOWED_EXTENSIONS = {'xlsx'}

class User(UserMixin):
    def __init__(self, id):
        self.id = id

@login_manager.user_loader
def load_user(user_id):
    return User(user_id)

class LoginForm(FlaskForm):
    username = StringField('Username', validators=[DataRequired()])
    password = PasswordField('Password', validators=[DataRequired()])
    submit = SubmitField('Login')
    
    
def convert_xls_to_xlsx_with_excel(input_file):
    pythoncom.CoInitialize()  # Initialize the COM library
    output_file = input_file + 'x'
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    try:
        # Open the file in Excel
        wb = excel.Workbooks.Open(input_file)

        # Save the file in the new format
        wb.SaveAs(output_file, FileFormat=51)  # 51 represents the .xlsx format

        wb.Close()
    finally:
        # Ensure Excel is closed
        excel.Application.Quit()
        pythoncom.CoUninitialize()  # Uninitialize the COM library
    
    time.sleep(5)
    
    return output_file


@main.route('/login', methods=['GET', 'POST'])
def login():
    form = LoginForm()
    if form.validate_on_submit():
        if form.username.data == "mer" and form.password.data == "eindregie":
            user = User(id=1)
            login_user(user)
            print("Form data:", form.data)
            print("Form errors:", form.errors)
            return redirect(url_for('main.index'))
        else:
            flash('Invalid username or password')
            
    return render_template('login.html', form=form)

@main.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('main.login'))

@main.route("/")
@login_required
def index():
    return render_template("index.html")

@main.route("/testing")
def testing():
    return render_template("testing.html")


# APPS -------------------------------------------------------------

@main.route("/app_01", methods=["GET", "POST"])
@login_required
def app_01():
    form = UploadForm()
    material_data = session.get('material_data', {})
    show_modal = False

    if form.validate_on_submit():
        
        file = form.file.data
        df = pd.read_excel(file)
        material_ids = df.iloc[:, 0].values.tolist()
        
        material_data = {
            "ids": material_ids,
            "replace_text": form.replace_text.data,
            "suffix": form.suffix.data,
            "material_type": form.material_type.data
        }
        
        session['material_data'] = material_data
        print(f"after session {material_data}")
        
        show_modal = True
        
        return render_template("app1.html", form=form, material_data=material_data, show_modal=show_modal)
        
    return render_template("app1.html", form=form, material_data=material_data, show_modal=show_modal)



@main.route("/app_02", methods=["GET", "POST"])
@login_required
def app_02():
    return render_template("app2.html")

@main.route("/app_03", methods=["GET", "POST"])
@login_required
def app_03():
    form = ExcelUploadForm()
    download_link = None  # Link to download the generated .sch file

    if form.validate_on_submit():
        if not os.path.exists(UPLOAD_FOLDER):
            os.makedirs(UPLOAD_FOLDER)

        file = form.file.data
        filename = secure_filename(file.filename)
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        file.save(filepath)
        
        # Process the file using the refactored script
        output_filepath = parse_excel_and_generate_playlist(filepath)
        
        if output_filepath:
            # File processed successfully, generate a download link
            download_link = url_for('main.download', filename=output_filepath.name)  # Using the name attribute of PurePath
            flash('File processed successfully!', 'success')
        else:
            # There was an error processing the file
            flash('An error occurred while processing the file. Please try again.', 'error')

    return render_template("app3.html", form=form, download_link=download_link)

@main.route('/uploads/<path:filename>', methods=["GET", "POST"])
def download(filename):
    return send_from_directory(UPLOAD_FOLDER, filename, as_attachment=True)


@main.route("/app_04", methods=["GET", "POST"])
@login_required
def app_04():
    color_mapping = {
        "yellow": "FFFFCD34",
        "red": "FFDC2626",
        "blue": "FF60A5FA",
        "orange": "FFFB923C",
    }
    file_uploaded = False  # or False based on your logic

    form = ColorForm()
    if form.validate_on_submit():
        # Hanlde the file upload
        if "file" not in request.files:
            return "No file part."
        file = request.files["file"]
        filename = secure_filename(file.filename)
        upload_dir = "uploads"
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)  # Create directory if it doesn't exist
        filepath = os.path.join(upload_dir, filename)
        print("Saving file as:", filepath)
        file.save(filepath)
        
        print("Absolute path of the file to convert:", os.path.abspath(filepath))

        time.sleep(2)
        
        if filepath.endswith('.xls'):
            print("Converting .xls to .xlsx")
            filepath = convert_xls_to_xlsx_with_excel(filepath)
        time.sleep(5)
        # Load the workbook
        book = load_workbook(filepath)

        # Define the target texts and fill colors
        target_texts_colors = {
            "VRT": PatternFill(
                start_color=color_mapping[form.programma.data],
                end_color=color_mapping[form.programma.data],
                fill_type="solid",
            ),
            "KT_2023WEEK": PatternFill(
                start_color=color_mapping[form.wrap.data],
                end_color=color_mapping[form.wrap.data],
                fill_type="solid",
            ),
            "V00": PatternFill(
                start_color=color_mapping[form.ondertiteling.data],
                end_color=color_mapping[form.ondertiteling.data],
                fill_type="solid",
            ),
        }

        # Define the font style
        font_style = Font(bold=True)

        # Apply conditional formatting
        for sheet in book.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        for target_text in target_texts_colors.keys():
                            if target_text in str(cell.value):
                                cell.fill = target_texts_colors[target_text]
                                cell.font = font_style
                                break

        # Save the workbook
        book.save(filepath)

        return redirect(url_for("main.download2", filename=filename))
    return render_template("app4.html", form=form, file_uploaded=file_uploaded)

@main.route("/download/<filename>")
def download2(filename):
    filepath = os.path.join(os.getcwd(), "uploads", filename)
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    else:
        return f"Error: File '{filename}' not found.", 404


@main.route("/app_05", methods=["GET", "POST"])
@login_required
def app_05():
    return render_template("app5.html")

# TOGGLE THEME -------------------------------------------------------------

@main.get("/toggle_theme")
def toggle_theme():
    current_theme = session.get('theme', 'emerald')
    new_theme = 'business' if current_theme == 'emerald' else 'emerald'
    session['theme'] = new_theme
    return redirect(request.args.get("current_page"))
